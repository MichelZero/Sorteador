import random
import openpyxl

def ler_nomes_de_xlsx(caminho_arquivo: str, coluna: str = 'Nome') -> list:
    """
    Lê nomes de uma coluna específica em um arquivo XLSX.

    Args:
        caminho_arquivo (str): O caminho para o arquivo XLSX.
        coluna (str): O nome da coluna que contém os nomes.

    Returns:
        list: Uma lista de nomes lidos do arquivo.
    """
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active
        nomes = []
        
        # Encontrar o índice da coluna
        header = [cell.value for cell in sheet[1]]
        try:
            coluna_idx = header.index(coluna)
        except ValueError:
            raise ValueError(f"Coluna '{coluna}' não encontrada no arquivo XLSX.")

        for row in sheet.iter_rows(min_row=2): # Começa da segunda linha para ignorar o cabeçalho
            if row[coluna_idx].value: # Verifica se a célula não está vazia
                nomes.append(str(row[coluna_idx].value).strip())
        return nomes
    except FileNotFoundError:
        raise FileNotFoundError(f"Arquivo XLSX não encontrado: {caminho_arquivo}")
    except Exception as e:
        raise Exception(f"Erro ao ler o arquivo XLSX: {e}")


class Sorteador:
    """
    Classe para realizar sorteio de nomes a partir de uma lista fornecida.
    """
    def __init__(self, lista: list):
        """
        Inicializa a classe Sorteador.

        Args:
            lista (list): Uma lista de nomes para o sorteio.
        """
        self.lista_nomes = lista
        self.nome_sorteado = None

    def sortear_nome(self) -> str:
        """
        Realiza o sorteio de um nome da lista.

        Returns:
            str: O nome sorteado.
        """
        if not self.lista_nomes:
            raise ValueError("A lista de nomes não pode estar vazia para realizar o sorteio.")
        self.nome_sorteado = random.choice(self.lista_nomes)
        return self.nome_sorteado

    def get_nome_sorteado(self) -> str:
        """
        Retorna o último nome sorteado.

        Returns:
            str: O nome sorteado. Retorna None se nenhum sorteio foi realizado.
        """
        return self.nome_sorteado

# Exemplo de uso da classe Sorteador
if __name__ == "__main__":
    import os
    from openpyxl import Workbook

    # Caminho do arquivo XLSX
    caminho_arquivo_xlsx = "data/lista_nomes.xlsx"
    
    # Criar o diretório 'data' se não existir
    os.makedirs(os.path.dirname(caminho_arquivo_xlsx), exist_ok=True)

    # Criar um arquivo XLSX de exemplo se não existir
    if not os.path.exists(caminho_arquivo_xlsx):
        print(f"Criando arquivo XLSX de exemplo: {caminho_arquivo_xlsx}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Participantes"
        ws['A1'] = "Nome"
        ws['A2'] = "Alice"
        ws['A3'] = "Beto"
        ws['A4'] = "Charles"
        ws['A5'] = "Davi"
        ws['A6'] = "Eva"
        wb.save(caminho_arquivo_xlsx)
    
    try:
        participantes = ler_nomes_de_xlsx(caminho_arquivo_xlsx)
        
        if not participantes:
            print("A lista de participantes está vazia no arquivo XLSX.")
        else:
            sorteador = Sorteador(participantes)

            print("Realizando sorteio...")
            ganhador = sorteador.sortear_nome()
            print(f"O nome sorteado foi: {ganhador}")

            # Você também pode obter o nome sorteado novamente
            print(f"Nome sorteado (via get_nome_sorteado): {sorteador.get_nome_sorteado()}")

    except (FileNotFoundError, ValueError, Exception) as e:
        print(f"\nErro: {e}")

    # Exemplo com lista vazia (deve gerar um erro)
    try:
        sorteador_vazio = Sorteador([])
        sorteador_vazio.sortear_nome()
    except ValueError as e:
        print(f"\nErro esperado ao tentar sortear com lista vazia: {e}")